import json
from urllib.request import urlopen
from django.shortcuts import render
from tienda.models import Producto
from django.views.generic import TemplateView, View
from openpyxl import Workbook
from django.http.response import HttpResponse
from tienda.utils import render_pdf
import requests
import json
from googletrans import Translator
from django.shortcuts import render

def tienda(request):
    product = Producto.objects.all()
    return render(request, 'tienda.html', {'productos': product})

def reporte(request):
    product = Producto.objects.all()
    return render(request, 'Listado.html', {'productos': product})
    

def api(request):
    url = "https://api.thecatapi.com/v1/breeds"
    headers = {
        'x-api-key': 'TU_CLAVE_API'  
    }

    response = requests.get(url, headers=headers)
    data = response.json()

    translator = Translator()
    diccionario = {}
    j = 0
    for i in data:
        material = translator.translate(i['name'], src='en', dest='es').text
        if material not in diccionario.values():
            diccionario[j] = material
            j += 1
    
    return render(request, 'api.html', {'diccionario': diccionario})



""" def api(request):
    url = "https://random-data-api.com/api/commerce/random_commerce?size=100"

    response = urlopen(url)
    data = json.load(response)
    diccionario = {}
    j = 0
    for i in data:
        if i ["material"] not in diccionario.values():
            diccionario[j]=i["material"]
            j +=1
    return render(request, 'api.html', {'diccionario': diccionario}) """


class ReporteProductoExcel(TemplateView):
    def get(self, request, *args, **Kwargs):
        producto = Producto.objects.all()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte de productos'
        ws['A1'] = 'Reporte de productos'
        ws.merge_cells('A1:C1')

        ws['A3'] = 'Producto'
        ws['B3'] = 'Precio'
        ws['C3'] = 'Imagen'
        ws['D3'] = 'Fecha de creación'
        cont = 4

        for product in producto:
            ws.cell(row=cont, column=1).value = product.nombre
            ws.cell(row=cont, column=2).value = product.precio
            ws.cell(row=cont, column=3).value = product.imagen.url
            ws.cell(row=cont, column=4).value = product.created
            cont += 1
        nombre_archivo = 'ReporteproductosExcel.xlsx'
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename={0}".format(nombre_archivo)
        response['content-Disposition'] = content
        wb.save(response)
        return response

def Listado(request):
    product = Producto.objects.all()
    return render(request, 'Listado.html', {'productos': product})


class ListaProductoPdf(View):
    def get(self, request, *args, **kwargs):
        productos = Producto.objects.all()
        data = {
            "productos": productos
        }
        pdf = render_pdf('tienda/listado.html', data)
        return HttpResponse(pdf, content_type='application/pdf')
